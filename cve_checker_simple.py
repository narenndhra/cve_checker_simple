#!/usr/bin/env python3
"""
═══════════════════════════════════════════════════════════════════════════
  CVE Nessus Coverage Checker  ·  Professional Dashboard Edition
═══════════════════════════════════════════════════════════════════════════

  Run:    python3 cve_nessus_pro.py
  Open:   http://localhost:5000

  HOW IT WORKS:
    1. Flask serves dashboard.html at localhost:5000
    2. You paste API keys in the UI (or edit config at top)
    3. First run builds a local CVE→plugin index (10-15 min with 30 workers)
    4. Subsequent CVE checks are instant against cached index
    5. Export results to Excel with color-coded sheets

  REQUIREMENTS:
    pip install flask requests openpyxl
═══════════════════════════════════════════════════════════════════════════
"""

import subprocess, sys
for _pkg in ["flask", "requests", "openpyxl"]:
    try:
        __import__(_pkg)
    except ImportError:
        print(f"[setup] Installing {_pkg}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", _pkg, "-q"])

import io, json, re, threading, time, socket
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, request, jsonify, send_file, Response

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 32 * 1024 * 1024

SCRIPT_DIR = Path(__file__).parent
DASH_PATH  = SCRIPT_DIR / "dashboard.html"

# Permanent local storage — never expires.
# Stored as a flat plugin database; the CVE index is rebuilt from it on load.
# Use "Update from Nessus" in the UI to fetch only NEW plugins.
CACHE_PATH = Path.home() / ".cve_nessus_cache.json"

# ────────────────────────────────────────────────────────────────────────────
#   GLOBAL STATE
# ────────────────────────────────────────────────────────────────────────────
_lock = threading.Lock()

state = {
    # connection
    "nessus_url":  "https://localhost:8834",
    "access_key":  "",
    "secret_key":  "",
    "connected":   False,
    "conn_msg":    "Not connected",
    "nessus_ver":  "",

    # index building
    "index_building":  False,
    "index_built":     False,
    "index_progress":  0,       # 0-100
    "index_status":    "",
    "index_total":     0,
    "index_done":      0,
    "index_cves":      0,
    "index_start_ts":  None,
    "index_eta":       0,
    "cve_index":       None,    # {CVE: [{id, name, cvss}, ...]}
    "abort_index":     False,

    # CVE scan
    "scanning":  False,
    "results":   [],
    "scan_total": 0,
    "scan_done":  0,

    # logs
    "log":     [],
    "log_seq": 0,
}

def add_log(level, msg):
    with _lock:
        state["log_seq"] += 1
        state["log"].append({
            "seq": state["log_seq"],
            "ts":  datetime.now().strftime("%H:%M:%S"),
            "lvl": level,
            "msg": msg,
        })
        if len(state["log"]) > 500:
            state["log"] = state["log"][-500:]


# ────────────────────────────────────────────────────────────────────────────
#   HELPERS
# ────────────────────────────────────────────────────────────────────────────
def build_headers():
    with _lock:
        ak, sk = state["access_key"], state["secret_key"]
    return {
        "X-ApiKeys": f"accessKey={ak}; secretKey={sk}",
        "Accept":    "application/json",
    }


def cvss_severity(s):
    try:
        v = float(s)
        if v >= 9.0: return "Critical"
        if v >= 7.0: return "High"
        if v >= 4.0: return "Medium"
        return "Low"
    except (TypeError, ValueError):
        return "N/A"


def extract_cves(text):
    seen, unique = set(), []
    for c in re.findall(r"CVE-\d{4}-\d{4,}", text, re.IGNORECASE):
        cu = c.upper()
        if cu not in seen:
            seen.add(cu)
            unique.append(cu)
    return unique


# ────────────────────────────────────────────────────────────────────────────
#   CONNECTION TEST
# ────────────────────────────────────────────────────────────────────────────
def test_connection(url, access_key, secret_key):
    headers = {
        "X-ApiKeys": f"accessKey={access_key}; secretKey={secret_key}",
        "Accept":    "application/json",
    }
    r = requests.get(
        f"{url.rstrip('/')}/server/properties",
        headers=headers, verify=False, timeout=15,
    )
    if r.status_code == 401:
        raise ValueError("API keys rejected (HTTP 401)")
    r.raise_for_status()
    return r.json().get("nessus_ui_version", "unknown")


# ────────────────────────────────────────────────────────────────────────────
#   BUILD CVE INDEX  (heavily parallel)
# ────────────────────────────────────────────────────────────────────────────
# ────────────────────────────────────────────────────────────────────────────
#   PERMANENT PLUGIN STORE  (plugin-keyed, never expires)
#
#   Storage format:
#     {
#       "url": "https://localhost:8834",
#       "version": 2,
#       "first_built_at": <unix ts>,
#       "last_updated_at": <unix ts>,
#       "plugins": {
#         "155998": {"name": "Apache Log4Shell RCE",
#                    "cves": ["CVE-2021-44228"],
#                    "cvss": "10.0"},
#         ...
#       }
#     }
#
#   The CVE→plugin index is built in memory from this on load.
#   Updates only fetch plugin IDs we don't already have → fast incremental sync.
# ────────────────────────────────────────────────────────────────────────────
def load_store(url):
    """Load the permanent plugin store. Returns plugins dict or None."""
    try:
        if not CACHE_PATH.exists():
            return None
        data = json.loads(CACHE_PATH.read_text(encoding="utf-8"))
        if data.get("url") != url:
            add_log("warn", f"Stored data is for a different URL — clearing")
            return None
        plugins = data.get("plugins", {})
        if not plugins:
            return None
        last = data.get("last_updated_at", 0)
        age_d = (time.time() - last) / 86400 if last else 0
        add_log("ok", f"Loaded {len(plugins):,} plugins from local store "
                      f"(last updated {age_d:.1f}d ago)")
        return plugins
    except Exception as e:
        add_log("warn", f"Store load failed: {e}")
        return None


def save_store(url, plugins, first_built_at=None):
    """Save the permanent plugin store."""
    try:
        CACHE_PATH.write_text(json.dumps({
            "url":            url,
            "version":        2,
            "first_built_at": first_built_at or time.time(),
            "last_updated_at": time.time(),
            "plugins":        plugins,
        }), encoding="utf-8")
        add_log("ok", f"Saved {len(plugins):,} plugins to {CACHE_PATH}")
    except Exception as e:
        add_log("warn", f"Store save failed: {e}")


def get_first_built(url):
    """Get the original build timestamp (so it persists across updates)."""
    try:
        if CACHE_PATH.exists():
            data = json.loads(CACHE_PATH.read_text(encoding="utf-8"))
            if data.get("url") == url:
                return data.get("first_built_at")
    except Exception:
        pass
    return None


def build_cve_index_from_plugins(plugins):
    """Rebuild the in-memory CVE→plugin lookup from the plugin store."""
    cve_index = {}
    for pid, pdata in plugins.items():
        for cid in pdata.get("cves", []):
            cve_index.setdefault(cid, []).append({
                "id":   pid,
                "name": pdata.get("name", ""),
                "cvss": pdata.get("cvss", "N/A"),
            })
    return cve_index


def _fetch_plugin_details(plugin_id, url, headers, session):
    """Fetch a single plugin's details. Returns (pid, cves_list, cvss, name) or None."""
    try:
        r = session.get(
            f"{url}/plugins/plugin/{plugin_id}",
            headers=headers, verify=False, timeout=15,
        )
        if not r.ok:
            return None
        pd = r.json()
        pname = pd.get("name", "")
        cvss  = "N/A"
        cves  = []
        for attr in pd.get("attributes", []):
            an = attr.get("attribute_name", "")
            av = attr.get("attribute_value", "")
            if an == "cve":
                for c in re.findall(r"CVE-\d{4}-\d+", av, re.I):
                    cves.append(c.upper())
            elif an == "cvss3_base_score":
                cvss = av
        return (plugin_id, cves, cvss, pname)
    except Exception:
        return None


def build_index_worker(url, concurrency, force_full=False):
    """
    Incremental index builder.

    Logic:
      1. Load existing plugin store from disk (if any).
      2. List all plugin IDs currently in Nessus.
      3. Compare → find NEW plugins (not in store) and STALE plugins (in store but
         no longer in Nessus, e.g. retired by Tenable).
      4. Only fetch the NEW plugin IDs.
      5. Merge with existing store, save, rebuild CVE index.

    First run: full build (~10 min for 290k plugins at 30 workers).
    Updates:   only new plugins (~30 sec for a few thousand new ones).

    If force_full=True, ignores existing store and re-fetches everything.
    """
    with _lock:
        state.update({
            "index_building":  True,
            "index_built":     False,
            "index_progress":  0,
            "index_status":    "Loading existing store...",
            "index_total":     0,
            "index_done":      0,
            "index_cves":      0,
            "index_start_ts":  time.time(),
            "abort_index":     False,
        })

    try:
        headers = build_headers()

        # Step 1 — load existing store
        existing_plugins = {} if force_full else (load_store(url) or {})
        first_built = get_first_built(url) if existing_plugins else None
        if existing_plugins:
            add_log("info", f"Existing store: {len(existing_plugins):,} plugins")
        else:
            add_log("info", "No existing store — full build")

        # Step 2 — list all plugin IDs in Nessus right now
        with _lock:
            state["index_status"] = "Fetching plugin families..."
        r = requests.get(f"{url}/plugins/families", headers=headers,
                         verify=False, timeout=30)
        r.raise_for_status()
        families = r.json().get("families", [])
        add_log("ok", f"Found {len(families)} plugin families")

        with _lock:
            state["index_status"] = f"Listing plugins from {len(families)} families..."

        all_current_ids = set()
        for fam in families:
            with _lock:
                if state.get("abort_index"):
                    state["index_building"] = False
                    add_log("warn", "Index build aborted")
                    return
            try:
                r2 = requests.get(f"{url}/plugins/families/{fam['id']}",
                                  headers=headers, verify=False, timeout=30)
                if not r2.ok:
                    continue
                for p in r2.json().get("plugins", []):
                    all_current_ids.add(str(p["id"]))
            except Exception:
                continue

        if not all_current_ids:
            raise RuntimeError("No plugins returned by Nessus")

        # Step 3 — diff against existing store
        existing_ids = set(existing_plugins.keys())
        new_ids      = all_current_ids - existing_ids
        retired_ids  = existing_ids - all_current_ids

        add_log("ok", f"Nessus has {len(all_current_ids):,} plugins · "
                      f"already stored: {len(existing_ids):,} · "
                      f"new to fetch: {len(new_ids):,} · "
                      f"retired: {len(retired_ids):,}")

        # Drop retired plugins from store
        for rid in retired_ids:
            existing_plugins.pop(rid, None)

        # Nothing new? We're done.
        if not new_ids:
            add_log("ok", "Store is already up-to-date — no new plugins to fetch")
            cve_index = build_cve_index_from_plugins(existing_plugins)
            save_store(url, existing_plugins, first_built_at=first_built)
            with _lock:
                state.update({
                    "cve_index":      cve_index,
                    "index_built":    True,
                    "index_building": False,
                    "index_progress": 100,
                    "index_status":   f"Up-to-date — {len(cve_index):,} CVEs indexed",
                })
            return

        # Step 4 — fetch ONLY new plugins (in parallel)
        total = len(new_ids)
        add_log("info", f"Fetching {total:,} new plugin(s) with {concurrency} workers...")
        with _lock:
            state["index_total"]  = total
            state["index_status"] = f"Fetching {total:,} new plugins..."

        plugins_db = dict(existing_plugins)   # start from existing
        done_count = 0
        _tls = threading.local()

        def worker_task(pid):
            if not hasattr(_tls, "session"):
                _tls.session = requests.Session()
                _tls.session.verify = False
            return _fetch_plugin_details(pid, url, headers, _tls.session)

        with ThreadPoolExecutor(max_workers=concurrency) as pool:
            futures = {pool.submit(worker_task, pid): pid for pid in new_ids}

            for fut in as_completed(futures):
                with _lock:
                    if state.get("abort_index"):
                        pool.shutdown(wait=False, cancel_futures=True)
                        # Save what we have so far (partial progress is preserved)
                        save_store(url, plugins_db, first_built_at=first_built)
                        add_log("warn", f"Aborted — saved {len(plugins_db):,} plugins")
                        state["index_building"] = False
                        return

                done_count += 1
                result = fut.result()
                if result:
                    pid, cves, cvss, pname = result
                    plugins_db[str(pid)] = {
                        "name": pname,
                        "cves": cves,
                        "cvss": cvss,
                    }

                if done_count % 25 == 0 or done_count == total:
                    pct = int(done_count / total * 100)
                    elapsed = time.time() - state["index_start_ts"]
                    rate = done_count / elapsed if elapsed > 0 else 0
                    eta  = int((total - done_count) / rate) if rate > 0 else 0
                    # Live CVE count from current plugins_db
                    cve_count = len({c for p in plugins_db.values() for c in p.get("cves", [])})
                    with _lock:
                        state["index_done"]     = done_count
                        state["index_progress"] = pct
                        state["index_cves"]     = cve_count
                        state["index_eta"]      = eta
                        state["index_status"]   = (
                            f"Fetching new plugins: {done_count:,}/{total:,} "
                            f"— total stored: {len(plugins_db):,}"
                        )

        # Step 5 — save merged store and rebuild CVE index
        elapsed = time.time() - state["index_start_ts"]
        add_log("ok", f"Index updated in {int(elapsed//60)}m{int(elapsed%60):02d}s")

        cve_index = build_cve_index_from_plugins(plugins_db)
        save_store(url, plugins_db, first_built_at=first_built)
        add_log("ok", f"Total: {len(plugins_db):,} plugins, {len(cve_index):,} unique CVEs")

        with _lock:
            state.update({
                "cve_index":      cve_index,
                "index_built":    True,
                "index_building": False,
                "index_progress": 100,
                "index_status":   f"Ready — {len(cve_index):,} CVEs from {len(plugins_db):,} plugins",
            })

    except Exception as e:
        add_log("err", f"Index build failed: {e}")
        with _lock:
            state.update({
                "index_building": False,
                "index_built":    False,
                "index_status":   f"Failed: {e}",
            })


# ────────────────────────────────────────────────────────────────────────────
#   CVE LOOKUP  (instant via index)
# ────────────────────────────────────────────────────────────────────────────
def check_cve(cve):
    with _lock:
        idx = state.get("cve_index")
    if idx is None:
        raise RuntimeError("Index not ready")
    entries = idx.get(cve.upper(), [])
    if not entries:
        return {
            "cve":          cve,
            "status":       "Not Covered",
            "plugin_count": 0,
            "plugin_ids":   "",
            "plugin_name":  "",
            "cvss":         "N/A",
        }
    return {
        "cve":          cve,
        "status":       "Covered",
        "plugin_count": len(entries),
        "plugin_ids":   ", ".join(str(e["id"]) for e in entries),
        "plugin_name":  entries[0]["name"],
        "cvss":         entries[0].get("cvss", "N/A"),
    }


def run_scan(cves):
    with _lock:
        state.update({
            "scanning":   True,
            "results":    [],
            "scan_total": len(cves),
            "scan_done":  0,
        })

    results = []
    for i, cve in enumerate(cves):
        try:
            r = check_cve(cve)
            results.append(r)
            with _lock:
                state["results"]   = list(results)
                state["scan_done"] = i + 1
        except Exception as e:
            add_log("err", f"Lookup failed for {cve}: {e}")

    with _lock:
        state["scanning"] = False

    cov = sum(1 for r in results if r["status"] == "Covered")
    nc  = sum(1 for r in results if r["status"] == "Not Covered")
    add_log("ok", f"Scan complete — {cov} covered, {nc} not covered")


# ────────────────────────────────────────────────────────────────────────────
#   FLASK ROUTES
# ────────────────────────────────────────────────────────────────────────────
@app.route("/")
def index_page():
    if not DASH_PATH.exists():
        return Response(
            "<h2>dashboard.html not found</h2>"
            "<p>Put <b>dashboard.html</b> in the same folder as this script.</p>",
            mimetype="text/html",
        )
    return Response(DASH_PATH.read_text(encoding="utf-8"), mimetype="text/html")


@app.route("/api/connect", methods=["POST"])
def api_connect():
    b = request.get_json(force=True) or {}
    url = (b.get("url") or "https://localhost:8834").rstrip("/")
    ak  = b.get("access_key", "").strip()
    sk  = b.get("secret_key", "").strip()

    if not ak or not sk:
        return jsonify({"ok": False, "msg": "API keys required"}), 400

    try:
        version = test_connection(url, ak, sk)
        with _lock:
            state.update({
                "nessus_url": url,
                "access_key": ak,
                "secret_key": sk,
                "connected":  True,
                "nessus_ver": version,
                "conn_msg":   f"Connected to Nessus {version}",
            })
        add_log("ok", f"Connected to Nessus {version} at {url}")

        # Check if we have a cached plugin store
        plugins = load_store(url)
        if plugins:
            cve_index = build_cve_index_from_plugins(plugins)
            with _lock:
                state.update({
                    "cve_index":      cve_index,
                    "index_built":    True,
                    "index_progress": 100,
                    "index_status":   f"Ready — {len(cve_index):,} CVEs from {len(plugins):,} plugins",
                })

        return jsonify({
            "ok":          True,
            "version":     version,
            "index_built": state["index_built"],
            "index_cves":  len(state["cve_index"]) if state["cve_index"] else 0,
        })
    except Exception as e:
        add_log("err", f"Connection failed: {e}")
        return jsonify({"ok": False, "msg": str(e)}), 401


@app.route("/api/build-index", methods=["POST"])
def api_build_index():
    with _lock:
        if not state["connected"]:
            return jsonify({"ok": False, "msg": "Connect first"}), 400
        if state["index_building"]:
            return jsonify({"ok": False, "msg": "Already building"}), 409

    b = request.get_json(force=True) or {}
    concurrency = max(1, min(100, int(b.get("concurrency", 30))))
    force_full  = bool(b.get("force_full", False))
    url = state["nessus_url"]

    threading.Thread(
        target=build_index_worker,
        args=(url, concurrency, force_full),
        daemon=True,
    ).start()

    mode = "Full rebuild" if force_full else "Incremental update"
    return jsonify({"ok": True, "msg": f"{mode} with {concurrency} workers"})


@app.route("/api/abort-index", methods=["POST"])
def api_abort_index():
    with _lock:
        state["abort_index"] = True
    add_log("warn", "Index build abort requested")
    return jsonify({"ok": True})


@app.route("/api/clear-cache", methods=["POST"])
def api_clear_cache():
    with _lock:
        state.update({
            "cve_index":      None,
            "index_built":    False,
            "index_progress": 0,
            "index_status":   "Cache cleared",
        })
    if CACHE_PATH.exists():
        CACHE_PATH.unlink()
    add_log("info", "Plugin index cache cleared")
    return jsonify({"ok": True})


@app.route("/api/upload-cves", methods=["POST"])
def api_upload_cves():
    if "file" not in request.files:
        return jsonify({"ok": False, "msg": "No file"}), 400
    f = request.files["file"]
    ext = (f.filename or "").rsplit(".", 1)[-1].lower()
    try:
        if ext in ("xlsx", "xls"):
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            text = "\n".join(
                " ".join(str(c) for c in row if c)
                for ws in wb.worksheets
                for row in ws.iter_rows(values_only=True)
            )
        else:
            text = f.read().decode("utf-8", errors="replace")
        cves = extract_cves(text)
        add_log("info", f"Parsed {f.filename}: {len(cves)} CVE IDs")
        return jsonify({"ok": True, "cves": cves, "count": len(cves), "filename": f.filename})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 500


@app.route("/api/scan", methods=["POST"])
def api_scan():
    with _lock:
        if state["scanning"]:
            return jsonify({"ok": False, "msg": "Scan in progress"}), 409
        if not state["index_built"]:
            return jsonify({"ok": False, "msg": "Index not built yet"}), 400

    b = request.get_json(force=True) or {}
    cves = b.get("cves", [])
    if not cves:
        return jsonify({"ok": False, "msg": "No CVEs provided"}), 400

    threading.Thread(target=run_scan, args=(cves,), daemon=True).start()
    return jsonify({"ok": True, "msg": f"Scanning {len(cves)} CVEs"})


@app.route("/api/status")
def api_status():
    since = int(request.args.get("since", 0))
    with _lock:
        new_logs = [l for l in state["log"] if l["seq"] > since]
        r = state["results"]
        return jsonify({
            "connected":      state["connected"],
            "nessus_ver":     state["nessus_ver"],
            "conn_msg":       state["conn_msg"],

            "index_building": state["index_building"],
            "index_built":    state["index_built"],
            "index_progress": state["index_progress"],
            "index_status":   state["index_status"],
            "index_total":    state["index_total"],
            "index_done":     state["index_done"],
            "index_cves":     state["index_cves"],
            "index_eta":      state["index_eta"],

            "scanning":       state["scanning"],
            "scan_total":     state["scan_total"],
            "scan_done":      state["scan_done"],
            "results":        r,
            "covered":        sum(1 for x in r if x["status"] == "Covered"),
            "not_covered":    sum(1 for x in r if x["status"] == "Not Covered"),

            "log":     new_logs,
            "log_seq": state["log_seq"],
        })


@app.route("/api/export")
def api_export():
    with _lock:
        results = list(state["results"])
        url     = state["nessus_url"]
    if not results:
        return jsonify({"ok": False, "msg": "No results to export"}), 400

    wb = openpyxl.Workbook()

    # Styles
    HDR_FILL  = PatternFill("solid", fgColor="1F2937")
    HDR_FONT  = Font(color="FFFFFF", bold=True, size=11)
    COV_FILL  = PatternFill("solid", fgColor="D1FAE5")
    NC_FILL   = PatternFill("solid", fgColor="FEE2E2")
    COV_FONT  = Font(color="065F46", bold=True)
    NC_FONT   = Font(color="991B1B", bold=True)
    CENTER    = Alignment(horizontal="center", vertical="center")
    LEFT      = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Sheet 1: CVE Coverage ────────────────────────────────────────
    ws = wb.active
    ws.title = "CVE Coverage"
    headers = ["#", "CVE ID", "Status", "CVSS v3", "Severity",
               "Plugin Count", "Plugin IDs", "Plugin Name"]
    ws.append(headers)
    for ci, _ in enumerate(headers, 1):
        c = ws.cell(1, ci)
        c.fill, c.font, c.alignment = HDR_FILL, HDR_FONT, CENTER

    for i, r in enumerate(results, 1):
        sev = cvss_severity(r.get("cvss", ""))
        ws.append([
            i, r["cve"], r["status"], r.get("cvss", "N/A"), sev,
            r.get("plugin_count", 0), r.get("plugin_ids", ""), r.get("plugin_name", ""),
        ])
        ri = ws.max_row
        fill = COV_FILL if r["status"] == "Covered" else NC_FILL
        font = COV_FONT if r["status"] == "Covered" else NC_FONT
        for ci in range(1, len(headers) + 1):
            c = ws.cell(ri, ci)
            c.fill = fill
            c.alignment = LEFT if ci in (7, 8) else CENTER
            if ci == 3:
                c.font = font

    widths = [5, 20, 16, 10, 12, 13, 50, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Summary ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    total = len(results)
    cov   = sum(1 for r in results if r["status"] == "Covered")
    nc    = sum(1 for r in results if r["status"] == "Not Covered")
    pct   = round(cov / total * 100, 1) if total else 0
    for row in [
        ["CVE Coverage Report", ""], ["", ""],
        ["Total CVEs", total],
        ["Covered", cov], ["Not Covered", nc],
        ["Coverage %", f"{pct}%"],
        ["", ""],
        ["Nessus URL", url],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]:
        ws2.append(row)
    ws2.cell(1, 1).font = Font(bold=True, size=14)
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 32
    ws2.cell(4, 1).fill = COV_FILL; ws2.cell(4, 2).fill = COV_FILL
    ws2.cell(5, 1).fill = NC_FILL;  ws2.cell(5, 2).fill = NC_FILL

    # ── Sheet 3: Not Covered ─────────────────────────────────────────
    ws3 = wb.create_sheet("Not Covered CVEs")
    ws3.append(["CVE ID", "CVSS v3", "Severity"])
    for ci in (1, 2, 3):
        c = ws3.cell(1, ci)
        c.fill, c.font, c.alignment = HDR_FILL, HDR_FONT, CENTER
    for r in results:
        if r["status"] == "Not Covered":
            ws3.append([r["cve"], r.get("cvss", "N/A"), cvss_severity(r.get("cvss", ""))])
    for i, w in enumerate([20, 12, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"
    if ws3.max_row > 1:
        ws3.auto_filter.ref = ws3.dimensions

    # ── Sheet 4: Covered ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Covered CVEs")
    ws4.append(["CVE ID", "CVSS v3", "Severity", "Plugin Count", "Plugin IDs", "Plugin Name"])
    for ci in range(1, 7):
        c = ws4.cell(1, ci)
        c.fill, c.font, c.alignment = HDR_FILL, HDR_FONT, CENTER
    for r in results:
        if r["status"] == "Covered":
            ws4.append([
                r["cve"], r.get("cvss", "N/A"), cvss_severity(r.get("cvss", "")),
                r.get("plugin_count", 0),
                r.get("plugin_ids", ""),
                r.get("plugin_name", ""),
            ])
    for i, w in enumerate([20, 12, 14, 13, 50, 55], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.freeze_panes = "A2"
    if ws4.max_row > 1:
        ws4.auto_filter.ref = ws4.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"CVE_Nessus_Coverage_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ────────────────────────────────────────────────────────────────────────────
#   STARTUP
# ────────────────────────────────────────────────────────────────────────────
def find_free_port(start=5000):
    for p in range(start, start + 20):
        with socket.socket() as s:
            try:
                s.bind(("127.0.0.1", p))
                return p
            except OSError:
                continue
    return start


if __name__ == "__main__":
    port = find_free_port(5000)
    url  = f"http://localhost:{port}"

    print()
    print("╔═══════════════════════════════════════════════════════════════╗")
    print("║  CVE Nessus Professional  ·  Dashboard Edition               ║")
    print("╠═══════════════════════════════════════════════════════════════╣")
    print(f"║  Dashboard  →  {url:<46} ║")
    print("║                                                               ║")
    print("║  First-time setup:                                            ║")
    print("║   1. Generate API keys in Nessus UI                          ║")
    print("║      (My Account → API Keys → Generate)                      ║")
    print("║   2. Paste them in the dashboard and Connect                 ║")
    print("║   3. Build the plugin index (one-time, ~10 min)              ║")
    print("║   4. Upload your CVE list and scan (instant)                 ║")
    print("║                                                               ║")
    print("║  Ctrl+C to stop.                                              ║")
    print("╚═══════════════════════════════════════════════════════════════╝")
    print()

    def _open_browser():
        time.sleep(1.2)
        try:
            import webbrowser
            webbrowser.open(url)
        except Exception:
            pass
    threading.Thread(target=_open_browser, daemon=True).start()

    app.run(host="127.0.0.1", port=port, debug=False, threaded=True)